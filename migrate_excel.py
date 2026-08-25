"""
One-time migration script: imports the existing 'Training Records' Excel sheet
(NO | Name of the Programme | From Date | To Date | Q | T/S | Names of the
Participants | Location | Training Cost | Training Hours | No Of people
attended | Total Hours) into the training_records Postgres table.

Usage:
    1. Put your secrets.toml at .streamlit/secrets.toml (see secrets.toml.example)
    2. pip install -r requirements.txt
    3. streamlit run app.py   # once, so init_db() creates the tables + first admin
    4. python migrate_excel.py path/to/HR_T___A.xlsx <admin_username>
"""

import sys
import openpyxl
from sqlalchemy import create_engine, text
import tomllib


def load_db_url():
    with open(".streamlit/secrets.toml", "rb") as f:
        secrets = tomllib.load(f)
    return secrets["postgres"]["url"]


def normalize_type(raw: str) -> str:
    raw = (raw or "").strip().lower()
    return "Soft Skill" if "soft" in raw else "Technical"


def normalize_quarter(raw: str):
    raw = (raw or "").strip().upper()
    return raw if raw in {"Q1", "Q2", "Q3", "Q4"} else None


def main(xlsx_path: str, admin_username: str):
    engine = create_engine(load_db_url())

    with engine.connect() as conn:
        admin = conn.execute(
            text("SELECT id FROM users WHERE username = :u"), {"u": admin_username}
        ).first()
    if not admin:
        print(f"User '{admin_username}' not found. Log in / sign up first, then re-run.")
        sys.exit(1)
    admin_id = admin[0]

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    inserted = 0
    with engine.begin() as conn:
        for row in ws.iter_rows(min_row=5, values_only=True):
            no = row[0]
            if no is None:
                continue
            (
                _no, programme_name, from_date, to_date, quarter, ttype,
                participants, location, cost, hours, attended, total_hours,
            ) = row[:12]

            if not programme_name or from_date is None:
                continue

            conn.execute(
                text(
                    """
                    INSERT INTO training_records
                        (programme_name, from_date, to_date, quarter, training_type,
                         location, participant_names, training_cost, training_hours,
                         participants_count, total_hours, created_by)
                    VALUES
                        (:programme_name, :from_date, :to_date, :quarter, :training_type,
                         :location, :participant_names, :training_cost, :training_hours,
                         :participants_count, :total_hours, :created_by)
                    """
                ),
                {
                    "programme_name": str(programme_name).strip(),
                    "from_date": from_date.date() if hasattr(from_date, "date") else from_date,
                    "to_date": (to_date.date() if hasattr(to_date, "date") else to_date) or from_date,
                    "quarter": normalize_quarter(quarter),
                    "training_type": normalize_type(ttype),
                    "location": (location or "HOF").strip(),
                    "participant_names": (participants or "").strip(),
                    "training_cost": float(cost or 0),
                    "training_hours": float(hours or 0),
                    "participants_count": int(attended or 0),
                    "total_hours": float(total_hours or (hours or 0) * (attended or 0)),
                    "created_by": admin_id,
                },
            )
            inserted += 1

    print(f"Imported {inserted} training records from {xlsx_path}.")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python migrate_excel.py <path_to_xlsx> <admin_username>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
